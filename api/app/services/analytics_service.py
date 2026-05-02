from sqlalchemy.orm import Session
from sqlalchemy import func, extract, case
from app.schemas.analytics import AnalyticsRead, MonthSales, AnnualSalesBreakdownRead
from app.db.models import PaymentSummary, PaymentTransaction, ClientOrder, Company, Client
from decimal import Decimal
from datetime import datetime, timezone
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def get_analytics(db: Session):
    current_date = datetime.now(timezone.utc)
    current_month = current_date.month
    current_year = current_date.year

    # Query 1: total uncollected balance
    total_balance = db.query(func.sum(PaymentSummary.remaining_balance)).join(
        ClientOrder, PaymentSummary.client_order_id == ClientOrder.id
    ).filter(
        PaymentSummary.isDeleted == False,
        ClientOrder.isDeleted == False
    ).scalar() or Decimal(0)

    # Query 2: completed + pending counts in one pass
    order_counts = db.query(
        func.sum(case((ClientOrder.isCompleted == True, 1), else_=0)).label("completed"),
        func.sum(case((ClientOrder.isCompleted == False, 1), else_=0)).label("pending"),
    ).filter(ClientOrder.isDeleted == False).one()

    # Query 3: monthly + annual sales in one pass
    sales = db.query(
        func.sum(case(
            (extract("month", PaymentTransaction.payment_date) == current_month,
             PaymentTransaction.paid_amount),
            else_=Decimal(0)
        )).label("monthly"),
        func.sum(PaymentTransaction.paid_amount).label("annual"),
    ).filter(
        extract("year", PaymentTransaction.payment_date) == current_year,
        PaymentTransaction.isDeleted == False
    ).one()

    return AnalyticsRead(
        total_balance=total_balance,
        total_completed_order=int(order_counts.completed or 0),
        total_pending_order=int(order_counts.pending or 0),
        monthly_sales=sales.monthly or Decimal(0),
        annual_sales=sales.annual or Decimal(0),
    )


def get_annual_sales_breakdown(db: Session, year_number: int = None):
    if year_number is None:
        year_number = datetime.now(timezone.utc).year

    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    # Single GROUP BY query instead of 12 separate queries
    month_col = extract("month", PaymentTransaction.payment_date).label("month")
    results = (
        db.query(month_col, func.sum(PaymentTransaction.paid_amount).label("sales"))
        .filter(
            extract("year", PaymentTransaction.payment_date) == year_number,
            PaymentTransaction.isDeleted == False
        )
        .group_by(month_col)
        .all()
    )
    sales_by_month = {int(row.month): row.sales for row in results}

    monthly_data = [
        MonthSales(
            month_number=m,
            month_name=month_names[m - 1],
            sales=sales_by_month.get(m, Decimal(0))
        )
        for m in range(1, 13)
    ]

    return AnnualSalesBreakdownRead(year_number=year_number, monthly_data=monthly_data)


def get_uncollected_balance(db: Session):
    from app.schemas.analytics import UncollectedBalanceItem

    # Query 1: all summary/order/client/company data in one join — no lazy loading
    rows = (
        db.query(
            PaymentSummary.id,
            PaymentSummary.remaining_balance,
            ClientOrder.order_date,
            ClientOrder.quantity,
            ClientOrder.price,
            Client.first_name,
            Client.last_name,
            Client.viber_number,
            Company.name,
        )
        .join(ClientOrder, PaymentSummary.client_order_id == ClientOrder.id)
        .join(Client, ClientOrder.client_id == Client.id)
        .join(Company, Client.company_id == Company.id)
        .filter(
            PaymentSummary.remaining_balance > 0,
            PaymentSummary.isDeleted == False,
            ClientOrder.isDeleted == False,
        )
        .all()
    )

    if not rows:
        return []

    # Query 2: fetch all relevant transactions in one query, build lookup dict
    summary_ids = [r[0] for r in rows]
    txns = (
        db.query(
            PaymentTransaction.payment_summary_id,
            PaymentTransaction.payment_number,
            PaymentTransaction.paid_amount,
        )
        .filter(
            PaymentTransaction.payment_summary_id.in_(summary_ids),
            PaymentTransaction.isDeleted == False,
        )
        .all()
    )
    pay_map: dict[int, dict[int, Decimal]] = {}
    for t_sid, t_num, t_amount in txns:
        pay_map.setdefault(t_sid, {})[t_num] = t_amount

    items = []
    for sid, balance, order_date, quantity, price, first, last, viber, company_name in rows:
        pays = pay_map.get(sid, {})
        items.append(UncollectedBalanceItem(
            company=company_name,
            name=f"{first} {last}",
            contact_number=viber,
            order_date=order_date,
            quantity=quantity,
            price=price,
            first_pay=pays.get(1, Decimal(0)),
            second_pay=pays.get(2, Decimal(0)),
            third_pay=pays.get(3, Decimal(0)),
            balance=balance,
        ))

    return items


def _style_header_row(ws, headers: list[str], fill_color: str = "550000"):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 20



def _write_sheet(ws, headers: list[str], rows):
    """Write header + rows to a sheet and auto-fit columns in a single pass."""
    _style_header_row(ws, headers)
    col_widths = [len(h) for h in headers]
    for row in rows:
        ws.append(row)
        for i, val in enumerate(row):
            if val is not None:
                col_widths[i] = min(max(col_widths[i], len(str(val))), 50)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w + 4


def generate_full_report(db: Session) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Sheet 1: Companies ---
    company_rows = [
        [c_id, name, addr or ""]
        for c_id, name, addr in (
            db.query(Company.id, Company.name, Company.address)
            .filter(Company.isDeleted == False)
            .order_by(Company.id)
            .all()
        )
    ]
    _write_sheet(wb.create_sheet("Companies"), ["ID", "Name", "Address"], company_rows)

    # --- Sheet 2: Clients ---
    client_rows = [
        [c_id, company_name, first, last, addr or "", viber, notes or ""]
        for c_id, company_name, first, last, addr, viber, notes in (
            db.query(
                Client.id, Company.name,
                Client.first_name, Client.last_name,
                Client.address, Client.viber_number, Client.notes
            )
            .join(Company, Client.company_id == Company.id)
            .filter(Client.isDeleted == False, Company.isDeleted == False)
            .order_by(Client.id)
            .all()
        )
    ]
    _write_sheet(
        wb.create_sheet("Clients"),
        ["ID", "Company", "First Name", "Last Name", "Address", "Viber Number", "Notes"],
        client_rows
    )

    # --- Sheet 3: All Orders ---
    order_cols = (
        db.query(
            ClientOrder.id, Company.name,
            Client.first_name, Client.last_name,
            ClientOrder.order_date, ClientOrder.model,
            ClientOrder.size, ClientOrder.material, ClientOrder.color, ClientOrder.mold,
            ClientOrder.heel_size, ClientOrder.heel_type,
            ClientOrder.has_platform, ClientOrder.has_slingback, ClientOrder.has_buckle,
            ClientOrder.quantity, ClientOrder.price,
            ClientOrder.isCompleted, ClientOrder.dateCompleted
        )
        .join(Client, ClientOrder.client_id == Client.id)
        .join(Company, Client.company_id == Company.id)
        .filter(ClientOrder.isDeleted == False)
        .order_by(ClientOrder.order_date.desc())
        .all()
    )
    order_rows = [
        [
            o_id, co_name, f"{first} {last}",
            od.strftime("%Y-%m-%d") if od else "",
            model, float(size), material, color, mold,
            heel_size, heel_type,
            "Yes" if platform else "No",
            "Yes" if slingback else "No",
            "Yes" if buckle else "No",
            qty, float(price),
            "Completed" if completed else "Pending",
            dc.strftime("%Y-%m-%d") if dc else ""
        ]
        for o_id, co_name, first, last, od, model, size, material, color, mold,
            heel_size, heel_type, platform, slingback, buckle, qty, price, completed, dc
        in order_cols
    ]
    _write_sheet(
        wb.create_sheet("All Orders"),
        [
            "Order ID", "Company", "Client Name", "Order Date",
            "Model", "Size", "Material", "Color", "Mold",
            "Heel Size", "Heel Type", "Platform", "Slingback", "Buckle",
            "Quantity", "Price", "Status", "Date Completed"
        ],
        order_rows
    )

    # --- Sheet 4: Payments ---
    txn_cols = (
        db.query(
            PaymentTransaction.id, ClientOrder.id,
            Company.name, Client.first_name, Client.last_name,
            PaymentTransaction.payment_number,
            PaymentTransaction.paid_amount, PaymentTransaction.payment_date
        )
        .join(PaymentSummary, PaymentTransaction.payment_summary_id == PaymentSummary.id)
        .join(ClientOrder, PaymentSummary.client_order_id == ClientOrder.id)
        .join(Client, ClientOrder.client_id == Client.id)
        .join(Company, Client.company_id == Company.id)
        .filter(PaymentTransaction.isDeleted == False, ClientOrder.isDeleted == False)
        .order_by(PaymentTransaction.payment_date.desc())
        .all()
    )
    txn_rows = [
        [
            tx_id, ord_id, co_name, f"{first} {last}",
            pay_num, float(amount),
            pd.strftime("%Y-%m-%d") if pd else ""
        ]
        for tx_id, ord_id, co_name, first, last, pay_num, amount, pd in txn_cols
    ]
    _write_sheet(
        wb.create_sheet("Payments"),
        ["Payment ID", "Order ID", "Company", "Client Name", "Payment #", "Paid Amount", "Payment Date"],
        txn_rows
    )

    # --- Sheet 5: Completed Orders ---
    completed_cols = (
        db.query(
            ClientOrder.id, Company.name,
            Client.first_name, Client.last_name,
            ClientOrder.order_date, ClientOrder.model,
            ClientOrder.size, ClientOrder.material, ClientOrder.color, ClientOrder.mold,
            ClientOrder.heel_size, ClientOrder.heel_type,
            ClientOrder.has_platform, ClientOrder.has_slingback, ClientOrder.has_buckle,
            ClientOrder.quantity, ClientOrder.price, ClientOrder.dateCompleted
        )
        .join(Client, ClientOrder.client_id == Client.id)
        .join(Company, Client.company_id == Company.id)
        .filter(ClientOrder.isDeleted == False, ClientOrder.isCompleted == True)
        .order_by(ClientOrder.dateCompleted.desc())
        .all()
    )
    completed_rows = [
        [
            o_id, co_name, f"{first} {last}",
            od.strftime("%Y-%m-%d") if od else "",
            model, float(size), material, color, mold,
            heel_size, heel_type,
            "Yes" if platform else "No",
            "Yes" if slingback else "No",
            "Yes" if buckle else "No",
            qty, float(price),
            dc.strftime("%Y-%m-%d") if dc else ""
        ]
        for o_id, co_name, first, last, od, model, size, material, color, mold,
            heel_size, heel_type, platform, slingback, buckle, qty, price, dc
        in completed_cols
    ]
    _write_sheet(
        wb.create_sheet("Completed Orders"),
        [
            "Order ID", "Company", "Client Name", "Order Date",
            "Model", "Size", "Material", "Color", "Mold",
            "Heel Size", "Heel Type", "Platform", "Slingback", "Buckle",
            "Quantity", "Price", "Date Completed"
        ],
        completed_rows
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()